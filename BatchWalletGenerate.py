from hdwallet import *
from hdwallet.cryptocurrencies import EthereumMainnet
from hdwallet.utils import generate_mnemonic
import os
import questionary
import openpyxl
from openpyxl.styles import Font, Alignment


def batch_wallet_generate():
    bip44_hdwallet: BIP44HDWallet = BIP44HDWallet(cryptocurrency=EthereumMainnet)

    mnemonic = generate_mnemonic(language="english", strength=128)

    bip44_hdwallet.from_mnemonic(mnemonic=mnemonic)
    words = bip44_hdwallet.mnemonic()

    if starts_with_0x == 'Yes':
        private_key = f'0x{bip44_hdwallet.private_key()}'
    else:
        private_key = f'{bip44_hdwallet.private_key()}'

    address = HDWallet(symbol='ETH').from_mnemonic(words).from_path("m/44'/60'/0'/0/0").p2pkh_address()

    wallet = [address, private_key, words]
    return wallet

def batch_wallet_restore(mnemo=''):
    bip44_hdwallet: BIP44HDWallet = BIP44HDWallet(cryptocurrency=EthereumMainnet)

    words = mnemo

    bip44_hdwallet.from_mnemonic(
        mnemonic=words, language="english"
    )

    if starts_with_0x == 'Yes':
        private_key = f'0x{bip44_hdwallet.private_key()}'
    else:
        private_key = f'{bip44_hdwallet.private_key()}'

    address = HDWallet(symbol='ETH').from_mnemonic(words).from_path("m/44'/60'/0'/0/0").p2pkh_address()

    wallet = [address, private_key, words]
    return wallet

def create_table(mode="1) Create wallets"):
    check = 0
    while True:
        filepath = os.getcwd() + f'\\wallets_{amount}' + ('' if check == 0 else f'_{check}') + '.xlsx'
        if not os.path.exists(filepath):
            break
        check += 1
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.title = 'WALLETS'
    ws.append(['ADDRESS', 'PRIVATE KEY', 'MNEMONIC'])

    for i in range(1, amount + 2):
        ws.row_dimensions[i].height = 25
    ws.column_dimensions['A'].width = 48
    ws.column_dimensions['B'].width = 72
    ws.column_dimensions['C'].width = 80

    if mode == "Create EVM wallets":
        for _ in range(amount):
            ws.append(batch_wallet_generate())

    if mode == "Restore EVM wallets from mnemonics":
        for mnemo in mnemo_list:
            ws.append(batch_wallet_restore(mnemo))

    for row in ws.iter_rows(min_row=1, max_row=amount + 1, min_col=1, max_col=3):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    for col in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=3):
        for cell in col:
            cell.font = Font(size=13,
                             bold=False)

    wb.save(filepath)
    print(f'\n[SUCCESS] Process finished | Path to file: {filepath}')

mode = questionary.select(
        "Select a mode to get started",
        choices=[
            "Create EVM wallets",
            "Restore EVM wallets from mnemonics",
        ],
        qmark="⚙️ ",
        pointer="✅ "
    ).ask()

if mode == "Create EVM wallets":

    starts_with_0x = questionary.select(
        'Do you want to add "0x" for privates? (0x9c43...)',
        choices=["Yes", "No"],
        qmark="⚙️ ",
        pointer="✅ "
    ).ask()

    amount = input('Type amount of wallets to create: ')
    while amount.isdigit() is False:
        amount = input('Type amount of wallets to create: ')
    amount = int(amount)

    create_table(mode)

elif mode == "Restore EVM wallets from mnemonics":

    starts_with_0x = questionary.select(
        'Do you want to add "0x" for privates? (0x9c43...)',
        choices=["Yes", "No"],
        qmark="⚙️ ",
        pointer="✅ "
    ).ask()

    with open("mnemonics.txt", "r") as f:
        mnemo_list = [row.strip() for row in f if row.strip()]

    amount = len(mnemo_list)
    print(f'[INFO] Amount of mnemonics in mnemonics.txt: {amount}')
    create_table(mode)






